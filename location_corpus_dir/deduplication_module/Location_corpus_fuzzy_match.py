import ast
import sys
import traceback

from rapidfuzz import process, fuzz
import pandas as pd
from openpyxl import load_workbook
from ast import literal_eval




class CityDeduplication():

    def __init__(self, location_corupus_file, bms_data_file):
        self.bms_data_file = bms_data_file
        self.location_corpus_df = pd.read_excel(location_corupus_file, sheet_name="Sheet1",
                                                converters={"PostalCode": int})

    @staticmethod
    def find_matches(matchThis, rows):
        FULL_MATCHING_THRESHOLD = 80
        PARTIAL_MATCHING_THRESHOLD = 90
        SORT_MATCHING_THRESHOLD = 90
        TOKEN_MATCHING_THRESHOLD = 90
        MAX_MATCHES = 1

        rows = [i.strip() for i in rows]
        rows = [i.lower() for i in rows]
        # rows.remove(matchThis)
        matches = process.extract(matchThis, rows, scorer=fuzz.ratio, score_cutoff=FULL_MATCHING_THRESHOLD,
                                  limit=MAX_MATCHES)
        if len(matches) == 0:
            matches = process.extract(matchThis, rows, scorer=fuzz.partial_ratio,
                                      score_cutoff=PARTIAL_MATCHING_THRESHOLD,
                                      limit=MAX_MATCHES)
            if len(matches) == 0:
                matches = process.extract(matchThis, rows, scorer=fuzz.token_set_ratio,
                                          score_cutoff=TOKEN_MATCHING_THRESHOLD, limit=MAX_MATCHES)
                if len(matches) == 0:
                    matches = process.extract(matchThis, rows, scorer=fuzz.token_sort_ratio,
                                              score_cutoff=SORT_MATCHING_THRESHOLD, limit=MAX_MATCHES)

        return matches if len(matches) > 0 else None

    def city_name_validation(self,city_col_num, state_col_num,status_col_num,output_col_num):

        workbook = load_workbook(self.bms_data_file)
        bms_data = workbook["Sheet1"]
        bms_data.cell(1, output_col_num).value = "Fuzzy Result"
        records_list = []
        for i in range(2, (bms_data.max_row + 1)):
            if bms_data.cell(i, status_col_num).value != "False":
                continue
            bms_city, bms_zip = bms_data.cell(i, city_col_num).value, int(bms_data.cell(i, state_col_num).value)
            try:
                filt = (self.location_corpus_df["PostalCode"] == bms_zip)
                corrected_city_list = self.location_corpus_df.loc[filt, 'CorrectedCities'].to_list()
                if not corrected_city_list:
                    matched_city_name = "no matching zip in LC"
                    bms_data.cell(i, output_col_num).value = matched_city_name
                else:
                    matched_city_name = CityDeduplication.find_matches(bms_city, literal_eval(corrected_city_list[0]))

                    bms_data.cell(i, output_col_num).value = matched_city_name[0][0] if matched_city_name is not None else "No match"
            except SyntaxError:
                records_list.append({"bms_row": i, "bms_zip": bms_zip})
            except:
                workbook.save(self.bms_data_file)
                traceback.print_exc()
                sys.exit()
        workbook.save(self.bms_data_file)
        print(records_list)


if __name__ == "__main__":
    location_corpus_file = "location_corpus_de_dup_more_than_2.xlsx"
    bms_data_file = "test_fuzzy.xlsx"
    CityDeduplication(location_corupus_file=location_corpus_file, bms_data_file=bms_data_file).city_name_validation(2,3,5,7)
