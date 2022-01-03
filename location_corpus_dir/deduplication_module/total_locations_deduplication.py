import traceback

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import sys
from openpyxl import load_workbook
import json

class TotalCorpusDeduplication():

    def __init__(self,driver_path):
        self.chrome_driver = webdriver.Chrome(driver_path)


    def wiki_scrape(self,location_to_search):
        self.chrome_driver.get("https://www.google.com")
        search = self.chrome_driver.find_element_by_name("q")
        search.clear()
        search.send_keys(location_to_search, Keys.RETURN)
        try:
            WebDriverWait(self.chrome_driver, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "gL9Hy")))

            google_filter = self.chrome_driver.find_element_by_css_selector("a[class='gL9Hy']")
            corrected_hyperlink_text = google_filter.text
            if self.state_for_validation.lower() in list(map(lambda x: x.lower(), corrected_hyperlink_text.split(" "))):
                google_filter.click()
        except:
            corrected_hyperlink_text = "None"
            print("No 'Showing results/did you mean'  found")

        try:
            WebDriverWait(self.chrome_driver, 19).until(
                EC.presence_of_element_located((By.CLASS_NAME, "LC20lb DKV0Md")))

            wiki_list = self.chrome_driver.find_elements_by_xpath('//a[contains(text(), "Wikipedia")]')
            wiki_text_list = list(map(lambda x:x.text,wiki_list))
            print(wiki_text_list)

        except:
            traceback.print_exc()




    def zip_validation(self, excel_name, sheetname):
        self.chrome_driver.get("https://www.google.com")
        workbook = load_workbook(excel_name)
        worksheet = workbook[sheetname]
        worksheet.cell(1, 4).value = "Status"
        worksheet.cell(1, 5).value = "google_recommended_city"
        worksheet.cell(1, 6).value = "google_recommended_state"
        worksheet.cell(1, 7).value = "google_recommended_county"
        worksheet.cell(1, 7).value = "corrected hyperlink text"
        temp_list = []

        for i in range(2, (worksheet.max_row + 1)):
            if i == 100:
                break
            try:
                # for index,row in location_df.iterrows():
                try:
                    state, city, zip_code = worksheet.cell(i, 2).value, worksheet.cell(i, 1).value, str(
                        worksheet.cell(i, 3).value)
                    self.state_for_validation = state
                except:
                    state, city, zip_code = "", "", ""
                location_to_search = " ".join((city, state))

                worksheet.cell(i, 4).value = ""
                worksheet.cell(i, 5).value = ""
                worksheet.cell(i, 6).value = ""
            except:
                workbook.save(excel_name)
                temp_list.append(i)
                sys.exit()
        workbook.save(excel_name)

        try:
            with open("errors.json", "w+") as f:
                json.dump({"errors": temp_list}, f)
        except:
            pass





if __name__ == "__main__":
    driver_path = r"../../chromedriver.exe"
    TotalCorpusDeduplication(driver_path).wiki_scrape("pa huntington")
