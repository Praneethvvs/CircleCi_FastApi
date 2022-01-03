import json
import traceback
from pprint import pprint
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from pprint import pprint
import time
import sys

class BmsLocationValidation():

    def __init__(self, driver_path):
        self.chrome_driver = webdriver.Chrome(driver_path)

    def wiki_scraping(self):
        pass

    def get_zips_bms(self, location_to_search):
        search = self.chrome_driver.find_element_by_name("q")
        search.clear()
        search.send_keys(location_to_search, Keys.RETURN)
        try:
            WebDriverWait(self.chrome_driver, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "gL9Hy")))

            google_filter = self.chrome_driver.find_element_by_css_selector("a[class='gL9Hy']")
            corrected_hyperlink_text = google_filter.text
            if self.state_for_validation.lower() in list(map(lambda x: x.lower(), corrected_hyperlink_text.split(" "))):
                google_filter.click()
        except:
            corrected_hyperlink_text = "None"
            print("No 'Showing results/did you mean'  found")

        try:
            WebDriverWait(self.chrome_driver, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "KKHQ8c")))
            zipcode_list_selector = self.chrome_driver.find_elements_by_css_selector("a[jscontroller='e8Ezlf']")
            zip_code_list = [i.get_attribute("data-entityname") for i in zipcode_list_selector]
            corrected_city_list = self.chrome_driver.find_elements_by_css_selector("span[class='Wkr6U z4P7Tc']")
            corrected_city = next(i.text for i in corrected_city_list if i.text != "Postcode")

            print("1", corrected_city, zip_code_list)
            return (corrected_city, zip_code_list), corrected_hyperlink_text

        except:
            try:
                WebDriverWait(self.chrome_driver, 1).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "FLP8od")))

                zipcode = self.chrome_driver.find_element_by_css_selector("a[class='FLP8od']").text
                corrected_city = self.chrome_driver.find_element_by_css_selector("span[class='GzssTd']").text
                print("2", corrected_city, zipcode)
                if corrected_city + zipcode + "2" == "2":
                    try:
                        WebDriverWait(self.chrome_driver, 1).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "IZ6rdc")))
                        zipcode = self.chrome_driver.find_element_by_css_selector("div[class='IZ6rdc']").text
                        corrected_city = ""
                        if not zipcode.isdigit():
                            return ("No result", "No result"), corrected_hyperlink_text
                    except:
                        return ("No result", "No result"), corrected_hyperlink_text
                return (corrected_city, zipcode), corrected_hyperlink_text

            except:
                print("no result")
                return ("No result", "No result"), corrected_hyperlink_text
                # try:
                #     WebDriverWait(self.chrome_driver, 2).until(
                #         EC.presence_of_element_located((By.CLASS_NAME, "IZ6rdc")))
                #     zip_code = self.chrome_driver.find_element_by_class_name("IZ6rdc").text
                #     corrected_city = self.chrome_driver.find_element_by_partial_link_text("wikipedia").text
                #     print(corrected_city,zip_code)
                #     return (corrected_city,zip_code)
                # except:
                #     return ("No result", "No result")

    def zip_validation(self, excel_name, sheetname):
        self.chrome_driver.get("https://www.google.com")
        workbook = load_workbook(excel_name)
        worksheet = workbook[sheetname]
        worksheet.cell(1, 4).value = "Status"
        worksheet.cell(1, 5).value = "google_recommended_city"
        worksheet.cell(1, 6).value = "corrected_hyperlink_text"
        worksheet.cell(1, 7).value = "google_recommended_state"
        temp_list = []

        for i in range(2, (worksheet.max_row + 1)):
            if i == 100:
                break
            try:
                # for index,row in location_df.iterrows():
                try:
                    state, city, zip_code = worksheet.cell(i, 2).value, worksheet.cell(i, 1).value, str(
                        worksheet.cell(i, 3).value)
                    self.state_for_validation = state
                except:
                    state, city, zip_code = "", "", ""
                location_to_search = " ".join((city, state, "zipcodes"))
                city_zip_result, corrected_hyperlink_text = self.get_zips_bms(location_to_search)
                rare_case_city_name = "No City"
                if corrected_hyperlink_text != "None":
                    split_correct_hyperlink_text = corrected_hyperlink_text.lower().split(" ")
                    rare_case_city_name = " ".join(
                        [i for i in split_correct_hyperlink_text if i.lower() not in ["zip", "code","zipcodes","codes", state.lower()]])

                if "No result" not in city_zip_result:
                    if isinstance(city_zip_result[1], list):
                        status = "True" if zip_code.zfill(5) in city_zip_result[1] else "False"
                    else:
                        status = "True" if zip_code.zfill(5) == city_zip_result[1] else "False"
                else:
                    status = "No result"
                worksheet.cell(i, 4).value = status
                worksheet.cell(i, 5).value = city_zip_result[0] if city_zip_result[0] != "" else rare_case_city_name.title()
                worksheet.cell(i, 6).value = corrected_hyperlink_text
            except:
                workbook.save(excel_name)
                temp_list.append(i)
                sys.exit()
        workbook.save(excel_name)

        try:
            with open("errors.json", "w+") as f:
                json.dump({"errors": temp_list}, f)
        except:
            pass

        #     location_df.loc[index,'status']=status
        # location_df.to_excel("duplicate_results.xlsx",index=False)


if __name__ == "__main__":
    excel_name = "remaining_zips_corpus.xlsx"
    driver_path = r"../chromedriver.exe"
    sheetname = "Sheet1"
    BmsLocationValidation(driver_path=driver_path).zip_validation(excel_name=excel_name,sheetname=sheetname)
