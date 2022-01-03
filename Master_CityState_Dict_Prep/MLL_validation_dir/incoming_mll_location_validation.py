import traceback

# import openpyxl
# import pandas as pd
#
#
# df = pd.read_excel("TEMP_FILE.xlsx")
# city_state = df["google_corrected_city"]+"-"+df["State"]
# city_state_list = city_state.to_list()
#
# workbook =openpyxl.load_workbook("processed_3201-4250.xlsx")
# worksheet = workbook["processed_3201-4250"]
# worksheet.cell(1,27).value = "Validations"
# for i in range(2, (worksheet.max_row + 1)):
#     try:
#         city,state = worksheet.cell(i,8).value,worksheet.cell(i,9).value
#         comb = city+"-"+state
#         if comb in city_state_list:
#             worksheet.cell(i, 27).value = "Yes"
#         else:
#             worksheet.cell(i, 27).value = "No"
#     except:
#         traceback.print_exc()
#         workbook.save("processed_3201-4250.xlsx")
# workbook.save("processed_3201-4250.xlsx")

import sys
import traceback
from tqdm import tqdm
import openpyxl
import pandas as pd
from rapidfuzz import process, fuzz


class ZipCorrection():

    @staticmethod
    def find_matches(matchThis, rows):
        FULL_MATCHING_THRESHOLD = 50
        PARTIAL_MATCHING_THRESHOLD = 90
        SORT_MATCHING_THRESHOLD = 50
        TOKEN_MATCHING_THRESHOLD = 90
        MAX_MATCHES = 1

        rows = [i.strip() for i in rows]
        rows = [i.lower() for i in rows]
        # rows.remove(matchThis)
        matches = process.extract(matchThis, rows, scorer=fuzz.ratio, score_cutoff=FULL_MATCHING_THRESHOLD,
                                  limit=MAX_MATCHES)

        return matches if len(matches) > 0 else None

    def validation(self):
        df = pd.read_excel("pulled_bms_locations.xlsx",sheet_name="Sheet1")
        df = df.apply(lambda x: x.astype(str).str.strip())
        city_state_dict = df.groupby("State")["City"].apply(list).to_dict()
        workbook = openpyxl.load_workbook("processed_Final_6100-6250_VALIDATED.xlsx")
        worksheet = workbook["Sheet1"]
        worksheet.cell(1, 27).value = "Match%"
        worksheet.cell(1, 28).value = "Validations"
        # 39781
        for i in tqdm(range(2, (worksheet.max_row + 1))):
            try:
                city, state = (worksheet.cell(i, 9).value).upper().strip(), (worksheet.cell(i, 10).value).upper().strip()
                if city_state_dict.get(state, None) != None:
                    if city in city_state_dict[state]:
                        worksheet.cell(i, 27).value = 100
                        worksheet.cell(i, 28).value = city
                    else:
                        result = ZipCorrection.find_matches(city, city_state_dict[state])
                        if result is not None:
                            worksheet.cell(i, 28).value = result[0][0]
                            worksheet.cell(i, 27).value = result[0][1]
                        else:
                            worksheet.cell(i, 28).value = "No match"
                            worksheet.cell(i, 27).value = -1
            except:
                traceback.print_exc()
                workbook.save("processed_Final_6100-6250_VALIDATED.xlsx")
                sys.exit()
        workbook.save("processed_Final_6100-6250_VALIDATED.xlsx")


ZipCorrection().validation()

