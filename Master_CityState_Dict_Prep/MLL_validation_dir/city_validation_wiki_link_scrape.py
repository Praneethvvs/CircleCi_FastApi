import sys
import time
import traceback
import pandas as pd
from rapidfuzz import process, fuzz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import json


class MasterLocationListCityCorrection():

    def pause(self):
        pause = input("Press Enter to continue")

    def __init__(self, driver_path):
        self.chrome_driver = webdriver.Chrome(driver_path)
        self.temp_input_save_dict = {}
        self.city_abb_dict = {
            "Alabama": "AL",
            "Alaska": "AK",
            "Arizona": "AZ",
            "Arkansas": "AR",
            "California": "CA",
            "Colorado": "CO",
            "Connecticut": "CT",
            "Delaware": "DE",
            "Florida": "FL",
            "Georgia": "GA",
            "Hawaii": "HI",
            "Idaho": "ID",
            "Illinois": "IL",
            "Indiana": "IN",
            "Iowa": "IA",
            "Kansas": "KS",
            "Kentucky": "KY",
            "Louisiana": "LA",
            "Maine": "ME",
            "Maryland": "MD",
            "Massachusetts": "MA",
            "Michigan": "MI",
            "Minnesota": "MN",
            "Mississippi": "MS",
            "Missouri": "MO",
            "Montana": "MT",
            "Nebraska": "NE",
            "Nevada": "NV",
            "New Hampshire": "NH",
            "New Jersey": "NJ",
            "New Mexico": "NM",
            "New York": "NY",
            "North Carolina": "NC",
            "North Dakota": "ND",
            "Ohio": "OH",
            "Oklahoma": "OK",
            "Oregon": "OR",
            "Pennsylvania": "PA",
            "Rhode Island": "RI",
            "South Carolina": "SC",
            "South Dakota": "SD",
            "Tennessee": "TN",
            "Texas": "TX",
            "Utah": "UT",
            "Vermont": "VT",
            "Virginia": "VA",
            "Washington": "WA",
            "West Virginia": "WV",
            "Wisconsin": "WI",
            "Wyoming": "WY"
        }

    @staticmethod
    def find_matches(matchThis, rows):
        FULL_MATCHING_THRESHOLD = 50
        PARTIAL_MATCHING_THRESHOLD = 90
        SORT_MATCHING_THRESHOLD = 50
        TOKEN_MATCHING_THRESHOLD = 90
        MAX_MATCHES = 1

        rows = [i.strip() for i in rows]
        rows = [i.lower() for i in rows]
        # rows.remove(matchThis)
        matches = process.extract(matchThis, rows, scorer=fuzz.ratio, score_cutoff=FULL_MATCHING_THRESHOLD,
                                  limit=MAX_MATCHES)
        if len(matches) == 0:
            matches = process.extract(matchThis, rows, scorer=fuzz.token_sort_ratio,
                                      score_cutoff=SORT_MATCHING_THRESHOLD,
                                          limit=MAX_MATCHES)
            # if len(matches) == 0:
            #     matches = process.extract(matchThis, rows, scorer=fuzz.token_set_ratio,
            #                               score_cutoff=TOKEN_MATCHING_THRESHOLD, limit=MAX_MATCHES)
            #     if len(matches) == 0:
            #         matches = process.extract(matchThis, rows, scorer=fuzz.partial_ratio,
            #                                   score_cutoff=SORT_MATCHING_THRESHOLD, limit=MAX_MATCHES)

        return matches if len(matches) > 0 else None

    def google_filter_click(self):
        try:
            WebDriverWait(self.chrome_driver, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "gL9Hy")))

            google_filter = self.chrome_driver.find_element_by_css_selector("a[class='gL9Hy']")
            corrected_hyperlink_text = google_filter.text
            if self.state_for_validation.lower() in list(
                    map(lambda x: x.lower().replace(",", ""), corrected_hyperlink_text.split(" "))):
                google_filter.click()
            else:
                corrected_hyperlink_text = "None"
        except:
            corrected_hyperlink_text = "None"


    def wiki_link_scrape(self, location_to_search):
        search = self.chrome_driver.find_element_by_name("q")
        search.clear()
        search.send_keys(location_to_search, Keys.RETURN)
        # self.google_filter_click()

        try:
            # WebDriverWait(self.chrome_driver, 2).until(
            #     EC.presence_of_element_located((By.CLASS_NAME, "LC20lb DKV0Md")))
            wiki_text_list = self.chrome_driver.find_elements_by_xpath("(//h3[contains(text(),' - Wikipedia')])")
            # self.pause()
            city_state_pair_list = []
            clean_text = lambda x: (x.replace("- Wikipedia", "").strip()).split(",")
            for value in wiki_text_list:
                text = value.text
                if text.count(",")==1:
                    result = clean_text(text)
                elif text.count(",")==2 and "county" in text.lower():
                    result = clean_text(text)
                    del result[1]
                else:
                    continue
                city_state_pair_list.append(result)

            # result_text_list = [(value.text).replace("- Wikipedia", "").strip() for value in wiki_text_list if
            #                     (value.text).count(",")==1]
            # city_state_pair_list = list(map(lambda x: x.split(","), result_text_list))

        except:
            city_state_pair_list = []
        return city_state_pair_list

    def city_correction(self, excel_name, sheet_name="Sheet1"):
        # df = pd.read_excel(excel_name,sheet_name=sheet_name)
        self.chrome_driver.get("https://www.google.com")
        workbook = load_workbook(excel_name)
        worksheet = workbook[sheet_name]
        worksheet.cell(1, 3).value = "google_corrected_city"
        worksheet.cell(1, 5).value = "State"
        worksheet.cell(1, 4).value = "Match%"

        for i in range(2, (worksheet.max_row + 1)):

            try:
                city, state = worksheet.cell(i, 1).value, worksheet.cell(i, 2).value
                self.state_for_validation = state.upper()
                key_exists = self.temp_input_save_dict.get(f"{city.lower()}-{state.lower()}", None)

                if city.isdigit() or state.isdigit():
                    continue
                if key_exists != None:
                    worksheet.cell(i, 3).value = key_exists[0]
                    worksheet.cell(i, 5).value = self.city_abb_dict[key_exists[1].title()]
                    worksheet.cell(i, 4).value = key_exists[2]

                else:
                    time.sleep(1)
                    city_state_pairs = self.wiki_link_scrape(f"{city} in {self.state_for_validation}, USA")

                    if not city_state_pairs:
                        continue
                    resultant_city_state_dict = {item[0].strip().lower(): item[1].strip().lower() for item in
                                                 city_state_pairs if self.city_abb_dict.get(item[1].strip().title(),"None")==self.state_for_validation}


                    fuzzy_matched_city = MasterLocationListCityCorrection.find_matches(city,resultant_city_state_dict.keys())

                    if fuzzy_matched_city != None:
                        result = [fuzzy_matched_city[0][0],resultant_city_state_dict[fuzzy_matched_city[0][0]],fuzzy_matched_city[0][1]]
                        worksheet.cell(i, 3).value = result[0]   #matched city
                        worksheet.cell(i, 5).value = self.city_abb_dict[result[1].title()]   # state
                        worksheet.cell(i, 4).value = result[2]   #match percentage
                        self.temp_input_save_dict.update({f"{city.lower()}-{state.lower()}":result})
                        print(resultant_city_state_dict, "--->", fuzzy_matched_city[0][1])

            except KeyboardInterrupt:
                traceback.print_exc()
                workbook.save(excel_name)
                sys.exit()
            except:
                traceback.print_exc()
                workbook.save(excel_name)
                sys.exit()
        workbook.save(excel_name)


if __name__ == "__main__":
    excel_name = "blanks.xlsx"
    driver_path = "../../chromedriver.exe"
    # MasterLocationListCityCorrection(driver_path=driver_path).city_correction(excel_name=excel_name,sheet_name="blanks")
    print(MasterLocationListCityCorrection.find_matches("miami",["miami beach"]))
