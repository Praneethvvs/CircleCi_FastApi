import sys
import time
import traceback
import pandas as pd
from rapidfuzz import process, fuzz
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import json


class MasterLocationListCityCorrection():

    def pause(self):
        input(f"{self.index})Press Enter to continue")

    def __init__(self, driver_path):
        self.chrome_driver = webdriver.Chrome(driver_path)
        self.temp_input_save_dict = {}
        self.city_abb_dict = {
            "Alabama": "AL",
            "Alaska": "AK",
            "Arizona": "AZ",
            "Arkansas": "AR",
            "California": "CA",
            "Colorado": "CO",
            "Connecticut": "CT",
            "Delaware": "DE",
            "Florida": "FL",
            "Georgia": "GA",
            "Hawaii": "HI",
            "Idaho": "ID",
            "Illinois": "IL",
            "Indiana": "IN",
            "Iowa": "IA",
            "Kansas": "KS",
            "Kentucky": "KY",
            "Louisiana": "LA",
            "Maine": "ME",
            "Maryland": "MD",
            "Massachusetts": "MA",
            "Michigan": "MI",
            "Minnesota": "MN",
            "Mississippi": "MS",
            "Missouri": "MO",
            "Montana": "MT",
            "Nebraska": "NE",
            "Nevada": "NV",
            "New Hampshire": "NH",
            "New Jersey": "NJ",
            "New Mexico": "NM",
            "New York": "NY",
            "North Carolina": "NC",
            "North Dakota": "ND",
            "Ohio": "OH",
            "Oklahoma": "OK",
            "Oregon": "OR",
            "Pennsylvania": "PA",
            "Rhode Island": "RI",
            "South Carolina": "SC",
            "South Dakota": "SD",
            "Tennessee": "TN",
            "Texas": "TX",
            "Utah": "UT",
            "Vermont": "VT",
            "Virginia": "VA",
            "Washington": "WA",
            "West Virginia": "WV",
            "Wisconsin": "WI",
            "Wyoming": "WY",
            "District Of Columbia":"DC",
            "Puerto Rico":"PR",
            "Virgin Islands":"VI"
        }


    def wiki_link_scrape(self, location_to_search):
        search = self.chrome_driver.find_element_by_name("q")
        search.clear()
        search.send_keys(location_to_search, Keys.RETURN)
        # self.google_filter_click()

        try:
            # WebDriverWait(self.chrome_driver, 2).until(
            #     EC.presence_of_element_located((By.CLASS_NAME, "LC20lb DKV0Md")))
            wiki_text_list = self.chrome_driver.find_elements_by_xpath("(//h3[contains(text(),' - Wikipedia')])")
            self.pause()
        except:
            pass



    def city_correction(self, excel_name, sheet_name="Sheet1"):
        # df = pd.read_excel(excel_name,sheet_name=sheet_name)
        temp_dict =  dict(zip(self.city_abb_dict.values(),self.city_abb_dict.keys()))
        self.chrome_driver.get("https://www.google.com")
        workbook = load_workbook(excel_name)
        worksheet = workbook[sheet_name]

        for i in range(2, (worksheet.max_row + 1)):
            self.index = i
            city, state = worksheet.cell(i, 1).value, worksheet.cell(i, 2).value
            try:

                self.state_for_validation = temp_dict[state.upper()]
                self.wiki_link_scrape(f"{city} in {self.state_for_validation}, USA")
            except:
                raise Exception




if __name__ == "__main__":
    excel_name = "Book1.xlsx"
    driver_path = "../../chromedriver.exe"
    MasterLocationListCityCorrection(driver_path=driver_path).city_correction(excel_name=excel_name)
    # print(MasterLocationListCityCorrection.find_matches("miami",["miami beach"]))
