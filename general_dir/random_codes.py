#
# from utils import *
#
# a = list(range(10000))
# b = list(range(50000,150000))
#
# def isthere(i,l2):
#     if i in l2:
#         return i
#     return False
#
# @profile
# def myfun(l1,l2):
#     newlist = []
#     for num in l1:
#         if isthere(num, l2):
#             newlist.append(num)
#     return newlist
#
import traceback

# import openpyxl
# import pandas as pd
#
#
# df = pd.read_excel("TEMP_FILE.xlsx")
# city_state = df["google_corrected_city"]+"-"+df["State"]
# city_state_list = city_state.to_list()
#
# workbook =openpyxl.load_workbook("processed_3201-4250.xlsx")
# worksheet = workbook["processed_3201-4250"]
# worksheet.cell(1,27).value = "Validations"
# for i in range(2, (worksheet.max_row + 1)):
#     try:
#         city,state = worksheet.cell(i,8).value,worksheet.cell(i,9).value
#         comb = city+"-"+state
#         if comb in city_state_list:
#             worksheet.cell(i, 27).value = "Yes"
#         else:
#             worksheet.cell(i, 27).value = "No"
#     except:
#         traceback.print_exc()
#         workbook.save("processed_3201-4250.xlsx")
# workbook.save("processed_3201-4250.xlsx")

import sys
import traceback
from tqdm import tqdm
import openpyxl
import pandas as pd
from rapidfuzz import process,fuzz


class ZipCorrection():

    @staticmethod
    def find_matches(matchThis, rows):
        FULL_MATCHING_THRESHOLD = 90
        PARTIAL_MATCHING_THRESHOLD = 90
        SORT_MATCHING_THRESHOLD = 50
        TOKEN_MATCHING_THRESHOLD = 90
        MAX_MATCHES = 1

        rows = [i.strip() for i in rows]
        rows = [i.lower() for i in rows]
        # rows.remove(matchThis)
        matches = process.extract(matchThis, rows, scorer=fuzz.ratio, score_cutoff=FULL_MATCHING_THRESHOLD,
                                  limit=MAX_MATCHES)

        return matches if len(matches) > 0 else None

    def validation(self):
        df = pd.read_excel("../SET_1_2_BMS.xlsx")
        df = df.apply(lambda x:x.astype(str).str.strip())
        city_state_dict = df.groupby("State")["google_corrected_city"].apply(list).to_dict()

        def myfun(city, state):
            if city_state_dict.get(state, None) != None:
                if city.strip().upper() in city_state_dict[state]:
                    return city
                else:
                    return False

        result_df = pd.read_excel("../processed_3201-4250.xlsx")

        result_df["new_col"] = result_df.apply(lambda x:myfun(x["City"],x["State"]),axis=1)
        result_df.to_excel("processed_3201-4250_result.xlsx")

ZipCorrection().validation()



